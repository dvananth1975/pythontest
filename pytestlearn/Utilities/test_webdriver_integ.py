import allure
import openpyxl
import pytest
from allure_commons.types import AttachmentType
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver import ChromeOptions, chrome
@pytest.fixture()
def log_on_failure(request,get_browser):
     yield
     item = request.node
     driver=get_browser
     if item.rep_call.failed:
        allure.attach(driver.get_screenshot_as_png(), name="dologin", attachment_type=AttachmentType.PNG)
def get_data():
    # return [
    #     ["anand@isos.com","Summerr@2022"],
    #     ["vijay600@isos.com","Summerr@2022"],
    #     ["vijay00@isos.com", "Summerr@2022"],
    # ]
    workbook = openpyxl.load_workbook("..//excel//testdata.xlsx")
    sheet = workbook["LoginTest"]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainlist = []
    for i in range(2, totalrows+1):
        dataList =[]
        for j in range(1,totalcols+1):
            data= sheet.cell(row=i,column=j).value
            dataList.insert(j,data)
        mainlist.insert(i,dataList)
    print(mainlist)
    return mainlist

@pytest.mark.usefixtures("log_on_failure")
@pytest.mark.parametrize("username,password",get_data())
def test_login(username,password,get_browser):
    driver = get_browser
    driver.find_element(By.XPATH ,"//*[@id=\"MainContent_LoginUser_txtUserName\"] ").send_keys(username)
    driver.find_element(By.XPATH,"//*[@id=\"MainContent_LoginUser_txtPassword\"]").send_keys(password)
    driver.find_element(By.XPATH,"//*[@id=\"onetrust-accept-btn-handler\"]").click()
    #assert 1 == 2
   # allure.attach(driver.get_screenshot_as_png(),name="dologin",attachment_type=AttachmentType.PNG)
