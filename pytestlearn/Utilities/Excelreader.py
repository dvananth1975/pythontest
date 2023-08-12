import openpyxl

def getrowcount(path,sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_row

def getcolcount(path,sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_column

def getcelldata(path,sheetname,rownum,colnum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum,column=colnum).value

path = "..//excel//testdata.xlsx"
sheetname = "LoginTest"
rows = getrowcount(path,sheetname)
cols = getcolcount(path,sheetname)

print( "Rows ", rows,"cols " , cols)

